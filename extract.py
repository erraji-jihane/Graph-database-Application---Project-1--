import os
import re
from pathlib import Path

import fitz
import cv2
import numpy as np
from PIL import Image
import pytesseract
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────── CONFIG ───────────────────────────────────
DPI = 450
OUTPUT_DIR = "output"
DEBUG_DIR = "debug_images"
SAVE_DEBUG = True

EXPECTED_BOXES = {1: 47, 2: 52}
PROCESS_ALL_PAGES_EXCEPT_LAST = True  # NEW: Skip last page

TESS_PATHS = [
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Users\PC\AppData\Local\Programs\Tesseract-OCR\tesseract.exe",
]

TSR_BLOCK = r"--oem 3 --psm 6"
TSR_SPARSE = r"--oem 3 --psm 11"

BOX_MIN_W, BOX_MIN_H = 90, 26
BOX_MAX_W, BOX_MAX_H = 1800, 1100

OUTLINE_COLOR = (0, 0, 255)    # red
LABEL_COLOR = (255, 0, 0)      # blue
OUTLINE_THICKNESS = 5

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(DEBUG_DIR, exist_ok=True)

def configure_tesseract():
    for p in TESS_PATHS:
        if os.path.exists(p):
            pytesseract.pytesseract.tesseract_cmd = p
            return
    print("⚠️ Tesseract executable not auto-found. Install it or update TESS_PATHS.")

# ══════════════════════════════════════════════════════════════════════
# STEP 1 — Find PDF
# ══════════════════════════════════════════════════════════════════════

def find_pdf() -> str:
    pdfs = sorted([f for f in Path(".").iterdir() if f.suffix.lower() == ".pdf"])
    if not pdfs:
        raise FileNotFoundError("No PDF found in current folder.")
    print(f"  ✅ Found PDF: {pdfs[0].name}")
    return str(pdfs[0])

# ══════════════════════════════════════════════════════════════════════
# STEP 2 — Render PDF pages to high-resolution images
# ══════════════════════════════════════════════════════════════════════

def pdf_to_images(pdf_path: str) -> list[tuple[int, np.ndarray]]:
    doc = fitz.open(pdf_path)
    pages = []

    for i, page in enumerate(doc):
        pix = page.get_pixmap(dpi=DPI, alpha=False)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        bgr = cv2.cvtColor(np.array(img), cv2.COLOR_RGB2BGR)
        pages.append((i + 1, bgr))

        if SAVE_DEBUG:
            cv2.imwrite(f"{DEBUG_DIR}/page_{i+1:02d}_original.png", bgr)

    doc.close()
    print(f"  Rendered {len(pages)} page(s) at {DPI} DPI")
    return pages

# ══════════════════════════════════════════════════════════════════════
# STEP 3 — Preprocess for OCR
# ══════════════════════════════════════════════════════════════════════

def preprocess_for_ocr(bgr: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)

    den = cv2.fastNlMeansDenoising(gray, h=10)

    blur = cv2.GaussianBlur(den, (0, 0), 3)
    sharp = cv2.addWeighted(den, 1.8, blur, -0.8, 0)

    bg = cv2.medianBlur(sharp, 31)
    norm = cv2.divide(sharp, bg, scale=255)

    adap = cv2.adaptiveThreshold(
        norm, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY, 31, 12
    )

    otsu = cv2.threshold(
        norm, 0, 255,
        cv2.THRESH_BINARY + cv2.THRESH_OTSU
    )[1]

    bw = cv2.bitwise_and(adap, otsu)

    k = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    bw = cv2.morphologyEx(bw, cv2.MORPH_OPEN, k, iterations=1)
    return bw

# ══════════════════════════════════════════════════════════════════════
# Box helpers
# ══════════════════════════════════════════════════════════════════════

def rect_area(b):
    return b[2] * b[3]

def rect_iou(a, b):
    ax, ay, aw, ah = a
    bx, by, bw, bh = b
    ax2, ay2 = ax + aw, ay + ah
    bx2, by2 = bx + bw, by + bh
    ix1, iy1 = max(ax, bx), max(ay, by)
    ix2, iy2 = min(ax2, bx2), min(ay2, by2)
    iw, ih = max(0, ix2 - ix1), max(0, iy2 - iy1)
    inter = iw * ih
    union = aw * ah + bw * bh - inter
    return inter / union if union else 0.0

def inside(a, b, pad=6):
    ax, ay, aw, ah = a
    bx, by, bw, bh = b
    return bx >= ax - pad and by >= ay - pad and bx + bw <= ax + aw + pad and by + bh <= ay + ah + pad

def sort_boxes(boxes):
    return sorted(boxes, key=lambda z: (z[1] // 40, z[0]))

def expand_box(x, y, w, h, W, H, pad=6):
    x1 = max(0, x - pad)
    y1 = max(0, y - pad)
    x2 = min(W, x + w + pad)
    y2 = min(H, y + h + pad)
    return (x1, y1, x2 - x1, y2 - y1)

def dedupe_boxes(boxes):
    boxes = sorted(boxes, key=rect_area, reverse=True)
    keep = []
    for b in boxes:
        bad = False
        for k in keep:
            if rect_iou(b, k) > 0.55 or inside(k, b):
                bad = True
                break
        if not bad:
            keep.append(b)
    return sort_boxes(keep)

# ══════════════════════════════════════════════════════════════════════
# STEP 4 — Detect course boxes
# ══════════════════════════════════════════════════════════════════════

def detect_boxes_pass(gray: np.ndarray, mode: int):
    H, W = gray.shape[:2]

    if mode == 1:
        th = cv2.adaptiveThreshold(
            gray, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY_INV, 21, 8
        )
        k = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5))
        th = cv2.morphologyEx(th, cv2.MORPH_CLOSE, k, iterations=2)

    elif mode == 2:
        th = cv2.adaptiveThreshold(
            gray, 255,
            cv2.ADAPTIVE_THRESH_MEAN_C,
            cv2.THRESH_BINARY_INV, 31, 10
        )
        k = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
        th = cv2.dilate(th, k, iterations=1)
        th = cv2.morphologyEx(th, cv2.MORPH_CLOSE, k, iterations=2)

    else:
        edges = cv2.Canny(gray, 40, 140)
        k = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
        th = cv2.dilate(edges, k, iterations=2)
        th = cv2.morphologyEx(th, cv2.MORPH_CLOSE, k, iterations=2)

    contours, _ = cv2.findContours(th, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)

    boxes = []
    for cnt in contours:
        peri = cv2.arcLength(cnt, True)
        approx = cv2.approxPolyDP(cnt, 0.03 * peri, True)
        x, y, w, h = cv2.boundingRect(cnt)

        if not (BOX_MIN_W <= w <= BOX_MAX_W and BOX_MIN_H <= h <= BOX_MAX_H):
            continue

        if w * h < 3500:
            continue

        ar = w / max(h, 1)
        if ar < 0.9 or ar > 14:
            continue

        if len(approx) < 4:
            continue

        boxes.append(expand_box(x, y, w, h, W, H, 6))

    return th, boxes

def find_boxes(bgr: np.ndarray, page_num: int) -> list[tuple[int, int, int, int]]:
    gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)
    gray = cv2.fastNlMeansDenoising(gray, h=8)

    all_boxes = []
    thresh_debug = []

    for mode in (1, 2, 3):
        th, boxes = detect_boxes_pass(gray, mode)
        thresh_debug.append((mode, th))
        all_boxes.extend(boxes)

    boxes = dedupe_boxes(all_boxes)

    expected = EXPECTED_BOXES.get(page_num)
    if expected and len(boxes) < expected:
        th = cv2.adaptiveThreshold(
            gray, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY_INV, 15, 5
        )
        k = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
        th = cv2.morphologyEx(th, cv2.MORPH_CLOSE, k, iterations=1)

        contours, _ = cv2.findContours(th, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
        H, W = gray.shape[:2]
        extra = []

        for cnt in contours:
            x, y, w, h = cv2.boundingRect(cnt)
            if 75 <= w <= 1500 and 22 <= h <= 800 and w * h >= 2600:
                extra.append(expand_box(x, y, w, h, W, H, 4))

        boxes = dedupe_boxes(boxes + extra)

        if SAVE_DEBUG:
            cv2.imwrite(f"{DEBUG_DIR}/page_{page_num:02d}_threshold_fallback.png", th)

    if SAVE_DEBUG:
        for mode, th in thresh_debug:
            cv2.imwrite(f"{DEBUG_DIR}/page_{page_num:02d}_threshold_mode{mode}.png", th)

        vis = bgr.copy()
        for i, (x, y, w, h) in enumerate(boxes, start=1):
            cv2.rectangle(vis, (x, y), (x + w, y + h), OUTLINE_COLOR, OUTLINE_THICKNESS)
            cv2.putText(
                vis, str(i), (x + 4, max(28, y - 8)),
                cv2.FONT_HERSHEY_SIMPLEX, 0.9, LABEL_COLOR, 3, cv2.LINE_AA
            )
        cv2.imwrite(f"{DEBUG_DIR}/page_{page_num:02d}_boxes_detected.png", vis)

    return boxes

# ══════════════════════════════════════════════════════════════════════
# STEP 5 — OCR isolated crop
# ══════════════════════════════════════════════════════════════════════

def ocr_box(bgr: np.ndarray, x: int, y: int, w: int, h: int, page_num: int, idx: int) -> str:
    PAD = 8
    H, W = bgr.shape[:2]
    crop = bgr[max(0, y-PAD):min(H, y+h+PAD), max(0, x-PAD):min(W, x+w+PAD)]

    bw = preprocess_for_ocr(crop)

    if bw.shape[0] < 90:
        scale = max(2, int(180 / max(1, bw.shape[0])))
        bw = cv2.resize(bw, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)

    if SAVE_DEBUG:
        cv2.imwrite(f"{DEBUG_DIR}/page_{page_num:02d}_box_{idx:03d}.png", bw)

    pil = Image.fromarray(bw)
    text = pytesseract.image_to_string(pil, config=TSR_BLOCK)
    if len(text.strip()) < 3:
        text = pytesseract.image_to_string(pil, config=TSR_SPARSE)
    return text

# ══════════════════════════════════════════════════════════════════════
# STEP 6 — Parse OCR text (UPDATED WITH CODE-BASED CREDITS)
# ══════════════════════════════════════════════════════════════════════

_CODE_RE = re.compile(r'\b([A-Z]{2,4})\s*(\d{4})\b')
_CREDIT_RE = re.compile(r'\b([1-9])\s*(?:SCH|credits?|cr\.?|hrs?)\b', re.I)
_CREDIT_P = re.compile(r'\(([1-9])\)')
_PREREQ_RE = re.compile(r'(?:pre-?req(?:uisite)?s?)\s*[:\-]?\s*(.+?)(?=\n|$)', re.I)
_NOISE_RE = re.compile(
    r'\b(prereq(?:uisite)?s?|total|semester|flowchart|catalog|page|note|fall|spring|summer|check|area|please)\b.*',
    re.I
)

def extract_credits_from_code(code: str) -> str:
    """Extract credits from course code - 2nd digit of the number part
    CSC1401 → 4 credits, CSC2302 → 3 credits"""
    digits = re.findall(r'\d', code)
    if len(digits) >= 2:
        return digits[1]  # 2nd digit: CSC1401→'4', CSC2302→'3'
    return "NONE"

def _clean(s: str) -> str:
    s = re.sub(r'[\n\r\t]+', ' ', s)
    s = re.sub(r'\s{2,}', ' ', s)
    return s.strip()

def parse_box_text(raw: str) -> dict | None:
    text = _clean(raw)
    if len(text) < 5:
        return None

    found = [a.upper() + b for a, b in _CODE_RE.findall(text)]
    if not found:
        return None

    code = found[0]
    rec = {
        "course_code": code,
        "course_title": "NONE",
        "course_credits": "NONE",
        "prerequisites": "NONE",
        "_raw": text[:700],
    }

    # PRIORITY 1: Extract credits from course code (2nd digit)
    code_credits = extract_credits_from_code(code)
    if code_credits != "NONE":
        rec["course_credits"] = code_credits
        print(f"  📊 Code-based credits: {code} → {code_credits}")

    # PRIORITY 2: Fallback to text patterns only if code extraction failed
    if rec["course_credits"] == "NONE":
        cm = _CREDIT_RE.search(text) or _CREDIT_P.search(text)
        if cm:
            rec["course_credits"] = cm.group(1)
            print(f"  📊 Text-based credits: {cm.group(0)}")

    pm = _PREREQ_RE.search(text)
    if pm:
        pr_raw = pm.group(1)
        pr_codes = [a.upper() + b for a, b in _CODE_RE.findall(pr_raw)]
        rec["prerequisites"] = ", ".join(pr_codes) if pr_codes else (_clean(pr_raw) or "NONE")

    # Clean title (remove code, credits, prereqs)
    title = text
    m = _CODE_RE.search(text)
    if m:
        title = title.replace(m.group(0), "", 1)
    cm = _CREDIT_RE.search(text) or _CREDIT_P.search(text)
    if cm:
        title = title[:cm.start()] + title[cm.end():]
    if pm:
        title = title[:pm.start()] + title[pm.end():]

    title = re.sub(r'\b[A-Z]{2,4}\s*\d{4}\b', ' ', title)
    title = re.sub(r'\b[1-9]\b', ' ', title)
    title = re.sub(r'\b\d+\s*SCH\b', ' ', title, flags=re.I)
    title = _NOISE_RE.sub('', title)
    title = re.sub(r'[^A-Za-z0-9 \-\&\/\:\,\']', ' ', title)
    title = _clean(title)

    if len(title) >= 3 and not title.isdigit():
        rec["course_title"] = title[:180]

    return rec

# ══════════════════════════════════════════════════════════════════════
# Extraction (UPDATED - SKIPS LAST PAGE)
# ══════════════════════════════════════════════════════════════════════

def extract_courses(pdf_path: str) -> list[dict]:
    pages = pdf_to_images(pdf_path)
    all_records = []

    # NEW: Skip last page if configured
    pages_to_process = pages[:-1] if PROCESS_ALL_PAGES_EXCEPT_LAST else pages
    total_pages = len(pages)
    pages_processed = len(pages_to_process)
    
    print(f"  📄 Processing {pages_processed}/{total_pages} pages (skipping last page)")
    
    for page_num, bgr in pages_to_process:
        print(f"\n  Page {page_num}:")
        boxes = find_boxes(bgr, page_num)
        print(f"    {len(boxes)} candidate boxes detected (expected {EXPECTED_BOXES.get(page_num, 'N/A')})")

        page_records = []
        for idx, (x, y, w, h) in enumerate(boxes, start=1):
            raw = ocr_box(bgr, x, y, w, h, page_num, idx)
            rec = parse_box_text(raw)
            if rec:
                page_records.append(rec)

        if SAVE_DEBUG:
            with open(f"{DEBUG_DIR}/page_{page_num:02d}_parsed.txt", "w", encoding="utf-8") as fh:
                fh.write(f"Detected boxes: {len(boxes)}\n")
                fh.write(f"Expected boxes: {EXPECTED_BOXES.get(page_num, 'N/A')}\n\n")
                for r in page_records:
                    fh.write(
                        f"{r['course_code']} | {r['course_title']} | "
                        f"{r['course_credits']} | {r['prerequisites']}\n"
                    )

        all_records.extend(page_records)

    best = {}
    for r in all_records:
        code = r["course_code"]
        if code not in best or len(r["course_title"]) > len(best[code]["course_title"]):
            best[code] = r

    print(f"\n  ✅ {len(best)} unique courses extracted")
    return list(best.values())

# ══════════════════════════════════════════════════════════════════════
# STEP 7 — Classification
# ══════════════════════════════════════════════════════════════════════

_MATH_PFX = ("MTH", "MAT")
_SCI_PFX = ("PHY", "BIO", "CHE", "STA")
_ENG_PFX = ("EGR", "ECE", "IEE", "MEE", "CEE")
_CS_PFX = ("CSC", "INF", "MIS", "CIS")
_GENED_PFX = ("ENG", "HUM", "SSC", "PSC", "PHI", "ARA", "ARB",
              "FRE", "FRN", "LIT", "COM", "PSY", "HIS", "ART",
              "ECO", "GEO", "SOC", "FYE", "FAS", "SLP")

_SS_W = ("social", "society", "econom", "geography", "psycho", "sociology", "political")
_HUM_W = ("human", "liter", "philos", "art", "histor", "islamic", "music",
          "drama", "dance", "painting", "aesthet", "creative")
_COM_W = ("english", "communicat", "writing", "speaking", "composition")
_AI_W = ("artificial intelligence", "machine learning", "deep learning",
         "neural network", "natural language", "computer vision", "nlp",
         "robotics", "data science", "reinforcement")

def _gened_category(title: str) -> str:
    t = title.lower()
    if any(w in t for w in _COM_W):
        return "GenEd_Communication"
    if any(w in t for w in _SS_W):
        return "GenEd_SocialSciences"
    if any(w in t for w in _HUM_W):
        return "GenEd_Humanities"
    return "GenEd"

def classify_courses(courses: list[dict]) -> dict[str, list[dict]]:
    groups = {
        "Mathematics": [],
        "ComputerScience": [],
        "ScienceEngineering": [],
        "GenEd": [],
        "Elective": [],
        "AI": [],
        "Minor": [],
    }

    for c in courses:
        code = c["course_code"]
        title_l = c.get("course_title", "").lower()
        raw_l = c.get("_raw", "").lower()

        if code.startswith(_MATH_PFX):
            c["category"] = "Mathematics"
            groups["Mathematics"].append(c)

        elif code.startswith(_SCI_PFX):
            c["category"] = "Science"
            groups["ScienceEngineering"].append(c)

        elif code.startswith(_ENG_PFX):
            c["category"] = "Engineering"
            groups["ScienceEngineering"].append(c)

        elif code.startswith(_CS_PFX):
            if "special" in raw_l or "specializ" in title_l:
                cat = "Specialization"
            elif "elective" in raw_l or "elective" in title_l:
                cat = "Computing elective"
            else:
                cat = "Required"
            c["category"] = cat
            groups["ComputerScience"].append(c)

        elif code.startswith(_GENED_PFX):
            c["category"] = _gened_category(c.get("course_title", ""))
            groups["GenEd"].append(c)

        else:
            c["category"] = "GenEd"
            groups["GenEd"].append(c)

    seen_e = set()
    for c in courses:
        if "elective" in c.get("_raw", "").lower() and c["course_code"] not in seen_e:
            groups["Elective"].append({**c, "category": "elective"})
            seen_e.add(c["course_code"])

    seen_ai = set()
    for c in courses:
        t = c.get("course_title", "").lower()
        r = c.get("_raw", "").lower()
        if any(k in t or k in r for k in _AI_W):
            if c["course_code"] not in seen_ai:
                groups["AI"].append({
                    **c,
                    "category": "Required" if "required" in r else "Optional"
                })
                seen_ai.add(c["course_code"])

    for i in range(1, 6):
        groups["Minor"].append({
            "course_code": f"course{i}",
            "course_title": "EMPTY",
            "course_credits": "NONE",
            "prerequisites": "EMPTY",
            "category": "Minor",
        })

    for k in groups:
        seen = {}
        for c in groups[k]:
            code = c["course_code"]
            if code not in seen or len(c.get("course_title", "")) > len(seen[code].get("course_title", "")):
                seen[code] = c
        groups[k] = list(seen.values())

    return groups

# ══════════════════════════════════════════════════════════════════════
# STEP 8 — Excel writer
# ══════════════════════════════════════════════════════════════════════

_HDR_FILL = PatternFill("solid", fgColor="1F497D")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
_ALT_FILL = PatternFill("solid", fgColor="DCE6F1")
_HEADERS = ["Course Code", "Course Title", "Course Credits", "Prerequisites", "Category"]
_WIDTHS = [16, 60, 15, 52, 26]

def write_excel(filename: str, rows: list[dict], title: str = None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (title or filename.replace(".xlsx", ""))[:31]

    ws.append(_HEADERS)
    for col in range(1, 6):
        c = ws.cell(1, col)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for i, r in enumerate(rows, start=2):
        ws.append([
            r.get("course_code", ""),
            r.get("course_title", "NONE"),
            r.get("course_credits", "NONE"),
            r.get("prerequisites", "NONE"),
            r.get("category", "NONE"),
        ])
        if i % 2 == 0:
            for col in range(1, 6):
                ws.cell(i, col).fill = _ALT_FILL
        for col in range(1, 6):
            ws.cell(i, col).alignment = Alignment(wrap_text=True, vertical="top")

    for idx, w in enumerate(_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"

    wb.save(os.path.join(OUTPUT_DIR, filename))
    print(f"  💾 {filename:<38} ({len(rows)} rows)")

# ══════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 60)
    print("  BSCSC — PDF → Images → OCR → Excel (with CODE-BASED CREDITS)")
    print("  🆕 SKIPS LAST PAGE automatically!")
    print("=" * 60)

    configure_tesseract()
    pdf_path = find_pdf()

    print("\n📄 Extracting courses via OCR ...")
    courses = extract_courses(pdf_path)

    print("\n🔀 Classifying ...")
    groups = classify_courses(courses)
    for sheet, lst in groups.items():
        print(f"     {sheet:<22} {len(lst):>3} rows")

    print("\n💾 Writing Excel files ...")
    write_excel("Mathematics.xlsx", groups["Mathematics"], "Mathematics")
    write_excel("ComputerScience.xlsx", groups["ComputerScience"], "ComputerScience")
    write_excel("ScienceEngineering.xlsx", groups["ScienceEngineering"], "ScienceEngineering")
    write_excel("GenEd.xlsx", groups["GenEd"], "GenEd")
    write_excel("Elective.xlsx", groups["Elective"], "Elective")
    write_excel("AI.xlsx", groups["AI"], "AI")
    write_excel("Minor.xlsx", groups["Minor"], "Minor")

    print("\n✅ Done!")
    print(f"   Excel files → {os.path.abspath(OUTPUT_DIR)}/")
    print(f"   Debug files → {os.path.abspath(DEBUG_DIR)}/")